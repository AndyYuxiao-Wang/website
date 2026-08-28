"""
Exports the "Demographics" sheet of demographics.xlsx (constituency-level
census/electoral demographics) into a clean JSON file for the web map's
demographic choropleth layer.

Only the well-formed demographic columns are kept - the workbook also
contains several derived/duplicate "tribe" columns (Muslim2, Left, Blues,
Column1..4, etc.) left over from other analysis, which are dropped here
since the map already has its own tribe pipeline (see 01_allocate_tribes.py)
and the "Electorate Segments" (2016) dataset.

Usage:
    py 06_export_demographics.py
"""
import openpyxl
import json
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BASE_DIR.parent
SOURCE = PROJECT_ROOT / "demographics.xlsx"
WEB_DATA = PROJECT_ROOT / "web" / "data"

# ---------------------------------------------------------------
# Source column -> (clean key, display label, group) map.
# Order here defines the export order; the group tag drives how the
# web panel buckets fields for the "Category" selector.
# ---------------------------------------------------------------

COLUMNS = [
    ("White",              "White",                  "White",                 "ethnicity"),
    ("White British",      "WhiteBritish",           "White British",         "ethnicity"),
    ("White Other",        "WhiteOther",             "White Other",           "ethnicity"),
    ("Asian",              "Asian",                  "Asian",                 "ethnicity"),
    ("Indian",             "Indian",                 "Indian",                "ethnicity"),
    ("Pakistani2",         "Pakistani",              "Pakistani",             "ethnicity"),
    ("Bangladeshi2",       "Bangladeshi",            "Bangladeshi",           "ethnicity"),
    ("Chinese",            "Chinese",                "Chinese",               "ethnicity"),
    ("Black",              "Black",                  "Black",                 "ethnicity"),
    ("Arab",               "Arab",                   "Arab",                  "ethnicity"),
    ("Mixed",              "Mixed",                  "Mixed",                 "ethnicity"),
    ("Other Ethnicity",    "OtherEthnicity",         "Other ethnicity",       "ethnicity"),

    ("Christian",          "Christian",              "Christian",             "religion"),
    ("No Religion",        "NoReligion",             "No religion",           "religion"),
    ("Muslim",             "Muslim",                 "Muslim",                "religion"),
    ("Hindi",              "Hindu",                  "Hindu",                 "religion"),
    ("Jewish",             "Jewish",                 "Jewish",                "religion"),
    ("Sikh",               "Sikh",                   "Sikh",                  "religion"),
    ("Bhuddist",           "Buddhist",               "Buddhist",              "religion"),

    ("0 to 17",            "Age0to17",               "Age 0-17",              "age"),
    ("18 to 24",           "Age18to24",              "Age 18-24",             "age"),
    ("65+",                "Age65plus",              "Age 65+",               "age"),

    ("No Qualifications",  "NoQualifications",       "No qualifications",     "education"),
    ("Degrees",            "Degrees",                "Degree educated",       "education"),

    ("PublicSector",              "PublicSector",            "Public sector worker",   "employment"),
    ("Managerial/Professional",   "ManagerialProfessional",  "Managerial/Professional", "employment"),
    ("Intermediate",              "Intermediate",            "Intermediate",           "employment"),
    ("Routine",                   "Routine",                 "Routine",                "employment"),

    ("SocialRent",         "SocialRent",             "Social rent",           "housing"),
    ("PrivateRent",        "PrivateRent",            "Private rent",          "housing"),
    ("Owner-Occupation",   "OwnerOccupation",        "Owner-occupied",        "housing"),

    ("Veterans",           "Veterans",               "Veterans",              "other"),
    ("LGBT",               "LGBT",                   "LGBT population",       "other"),
]

GROUP_LABELS = {
    "ethnicity": "Ethnicity",
    "religion": "Religion",
    "age": "Age",
    "education": "Education",
    "employment": "Employment",
    "housing": "Housing tenure",
    "other": "Other",
}


def load_descriptions(wb):
    """Seat -> free-text writeup, from the Buildip sheet's "Column22" column.
    Still being filled in by hand (partial coverage as of writing) - seats
    without one just get None, and the panel hides the box for those."""
    ws = wb["Buildip"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_index = {name: i for i, name in enumerate(header)}
    seat_i = col_index["Seat"]
    desc_i = col_index["Column22"]

    descriptions = {}
    for row in rows[1:]:
        seat = row[seat_i]
        if not seat:
            continue
        desc = row[desc_i]
        if isinstance(desc, str) and desc.strip():
            descriptions[seat] = desc.strip()
    return descriptions


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    descriptions = load_descriptions(wb)

    ws = wb["Demographics"]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_index = {name: i for i, name in enumerate(header)}

    missing = [src for src, _, _, _ in COLUMNS if src not in col_index]
    if missing:
        raise SystemExit(f"Columns missing from Demographics sheet: {missing}")

    seat_i = col_index["Seat"]
    region_i = col_index.get("Region")
    county_i = col_index.get("County")

    records = []

    for row in rows[1:]:
        seat = row[seat_i]
        if not seat:
            continue

        record = {"Name": seat}
        if region_i is not None:
            record["Region"] = row[region_i]
        if county_i is not None:
            record["County"] = row[county_i]
        record["Description"] = descriptions.get(seat)

        for src, key, _label, _group in COLUMNS:
            value = row[col_index[src]]
            record[key] = round(value * 100, 3) if isinstance(value, (int, float)) else None

        records.append(record)

    WEB_DATA.mkdir(parents=True, exist_ok=True)

    with open(WEB_DATA / "demographics.json", "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2)

    # Grouping metadata the JS panel uses to build the Category/Field
    # selectors, kept in the same file the pipeline writes so it can't
    # silently drift from the actual exported keys.
    fields_meta = [
        {"key": key, "label": label, "group": group}
        for _src, key, label, group in COLUMNS
    ]
    manifest = {
        "groups": [{"key": k, "label": v} for k, v in GROUP_LABELS.items()],
        "fields": fields_meta,
    }
    with open(WEB_DATA / "demographics_manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    with_desc = sum(1 for r in records if r["Description"])
    print(f"Wrote {len(records)} constituencies to {WEB_DATA / 'demographics.json'} ({with_desc} with a description)")
    print(f"Wrote field manifest to {WEB_DATA / 'demographics_manifest.json'}")


if __name__ == "__main__":
    main()
