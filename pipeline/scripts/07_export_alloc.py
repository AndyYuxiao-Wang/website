"""
Exports the tribe-level seat allocation (englandAlloc.csv / scotlandAlloc.csv
/ walesAlloc.csv, produced by 01_allocate_tribes.py) plus predictionTable.xlsx
as a single JSON the web app can load client-side, for the "Custom" tab's
interactive what-if predictor (web/js/customPredictor.js). Nothing else reads
this file - it isn't consumed by any other pipeline stage.

Re-run this whenever the Alloc files or predictionTable.xlsx change (same
convention as 06_export_demographics.py).

Usage:
    py 07_export_alloc.py
"""
import json
import openpyxl
import pandas as pd
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = BASE_DIR.parent
INTERMEDIATE = BASE_DIR / "data" / "intermediate"
RAW = BASE_DIR / "data" / "raw"
PREDICTION_TABLE = PROJECT_ROOT / "predictionTable.xlsx"
WEB_DATA = PROJECT_ROOT / "web" / "data"

# Alloc party columns -> the pipeline's canonical party names (matches
# 02_project_flows.py's flow-matrix party set, minus Restore, which never
# appears in the baseline - only as a possible destination).
ALLOC_PARTIES = ["Labour", "Conservative", "Reform", "LibDem", "Green", "Oth", "SNP", "Plaid"]
PIPELINE_PARTIES = ALLOC_PARTIES + ["Restore"]
SEGMENTS = ["Muslim", "Left", "Progressives", "Average", "Liberal", "Blues", "Reforms"]

NATIONS = {
    "england": "englandAlloc.csv",
    "scotland": "scotlandAlloc.csv",
    "wales": "walesAlloc.csv",
}

FLOWS_FILES = {
    "england": "englandFlows.xlsx",
    "scotland": "scotlandFlows.xlsx",
    "wales": "walesFlows.xlsx",
}

PREV_COLS = [
    "prev_Electorate", "prev_LAB", "prev_CON", "prev_REF", "prev_LD",
    "prev_GRN", "prev_OTHER", "prev_SNP", "prev_PLAID", "prev_OTH",
    "prev_UKIP", "prev_BNP", "prev_RES", "prev_TOTAL",
]

# predictionTable.xlsx spells parties/segments slightly differently from the
# pipeline's own canonical names (e.g. "Lib Dem" vs "LibDem", "Other" vs
# "Oth", segment "Reform" vs tribe "Reforms") - normalise on export so the
# JS side only ever has to deal with one naming scheme.
PARTY_NAME_MAP = {
    "Reform": "Reform",
    "Conservative": "Conservative",
    "Labour": "Labour",
    "Other": "Oth",
    "Oth": "Oth",
    "SNP": "SNP",
    "Plaid": "Plaid",
    "Green": "Green",
    "Lib Dem": "LibDem",
    "LibDem": "LibDem",
    "Restore": "Restore",
    # A handful of short party-ticker aliases show up inconsistently in
    # some hand-maintained sheets (e.g. Winners has "GRN" for Green).
    "GRN": "Green",
    "LAB": "Labour",
    "CON": "Conservative",
    "LD": "LibDem",
    "REF": "Reform",
}


def normalize_party(name):
    if name not in PARTY_NAME_MAP:
        raise SystemExit(f"Unrecognised party name {name!r} in a Tactical.xlsx/LocalFlows.xlsx sheet")
    return PARTY_NAME_MAP[name]

SEGMENT_NAME_MAP = {
    "Muslim": "Muslim",
    "Reform": "Reforms",
    "Average": "Average",
    "Progressive": "Progressives",
    "Liberal": "Liberal",
    "Left": "Left",
}


def main():
    vote_totals = pd.read_excel(RAW / "Tactical.xlsx", sheet_name="VoteTotals")
    vote_totals = vote_totals.set_index("Seat")

    output = {"nations": {}, "predictionTable": []}

    for nation, fname in NATIONS.items():
        alloc = pd.read_csv(INTERMEDIATE / fname)
        # Not every nation's ballot has every party (e.g. Wales has no SNP
        # column at all) - treat any missing party column as always 0.
        present_parties = [p for p in ALLOC_PARTIES if p in alloc.columns]

        seats = {}
        for _, row in alloc.iterrows():
            seat = row["Seat"]
            tribe = row["Tribe"]
            seats.setdefault(seat, {})[tribe] = {
                p: (float(row[p]) if p in present_parties else 0.0)
                for p in ALLOC_PARTIES
            }

        seat_records = []
        total_votes = 0.0
        for seat, segments in seats.items():
            vt = vote_totals.loc[seat]
            prev = {c: (None if pd.isna(vt[c]) else float(vt[c])) for c in PREV_COLS}
            seat_records.append({
                "Seat": seat,
                "segments": segments,
                **prev,
            })
            total_votes += prev["prev_TOTAL"] or 0.0

        output["nations"][nation] = {
            "seats": seat_records,
            "totalVotes": total_votes,
        }

        print(f"{nation}: {len(seat_records)} seats, {total_votes:.0f} total votes")

    wb = openpyxl.load_workbook(PREDICTION_TABLE, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for r in rows:
        if not r or r[0] is None:
            continue
        party_a, party_b, segment = r[0], r[1], r[2]
        for name in (party_a, party_b):
            if name not in PARTY_NAME_MAP:
                raise SystemExit(f"predictionTable.xlsx: unrecognised party name {name!r}")
        if segment not in SEGMENT_NAME_MAP:
            raise SystemExit(f"predictionTable.xlsx: unrecognised segment name {segment!r}")
        output["predictionTable"].append({
            "partyA": PARTY_NAME_MAP[party_a],
            "partyB": PARTY_NAME_MAP[party_b],
            "segment": SEGMENT_NAME_MAP[segment],
        })

    print(f"predictionTable: {len(output['predictionTable'])} rows")

    # ---- Flow matrices (englandFlows.xlsx etc, one sheet per segment) ----
    # These are the same hand-tuned party->party transition matrices
    # 02_project_flows.py uses for the real "Prediction" tab - the Custom
    # Predictor uses them as its starting point, then layers its own
    # churn-table adjustments on top to bridge to the user's targets,
    # rather than starting from "nobody moves".
    output["flowMatrices"] = {}
    for nation, fname in FLOWS_FILES.items():
        wb_flows = openpyxl.load_workbook(RAW / fname, data_only=True)
        nation_matrices = {}
        for seg in SEGMENTS:
            ws = wb_flows[seg]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_index = {name: i for i, name in enumerate(header) if name in PIPELINE_PARTIES}

            seg_matrix = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                origin = row[0]
                if origin not in PIPELINE_PARTIES:
                    continue
                seg_matrix[origin] = {}
                for dest in PIPELINE_PARTIES:
                    idx = col_index.get(dest)
                    val = row[idx] if idx is not None and idx < len(row) else None
                    seg_matrix[origin][dest] = float(val) if val is not None else 0.0

            # Any pipeline party never listed as an origin row in this sheet
            # (e.g. Restore, or a nation-exclusive party with no flows
            # defined) just stays 100% put by default.
            for origin in PIPELINE_PARTIES:
                if origin not in seg_matrix:
                    seg_matrix[origin] = {dest: (100.0 if dest == origin else 0.0) for dest in PIPELINE_PARTIES}

            nation_matrices[seg] = seg_matrix
        output["flowMatrices"][nation] = nation_matrices
        print(f"{nation} flow matrices: {len(nation_matrices)} segments")

    # ---- Tactical voting inputs (Tactical.xlsx) --------------------------
    # Same sheets 04_tactical_voting.py reads - exported so the Custom
    # Predictor can run the identical tactical-voting pass (tiered
    # eligibility by gap-to-leader, TVMatrix ranking order, TVPCT/
    # willingness-weighted cascade) on its own projected result, not just
    # the fixed-model one.
    tactical_wb = openpyxl.load_workbook(RAW / "Tactical.xlsx", data_only=True)

    # TVPCT is now a matrix: column A = donor party, column B = that
    # donor's flat "TVPct" (% of its own voters willing to consider
    # tactical voting at all - unchanged role), columns C onward = that
    # donor's per-recipient willingness % (how appealing each specific
    # recipient is to its tactical-considering voters - not a split, these
    # don't sum to 100 across a row).
    tv_ws = tactical_wb["TVPCT"]
    tv_header = [c.value for c in next(tv_ws.iter_rows(min_row=1, max_row=1))]
    tv_col_index = {normalize_party(name): i for i, name in enumerate(tv_header) if name and name != "TVPct"}
    output["tvPct"] = {}
    output["tvWillingness"] = {}
    for row in tv_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        donor = normalize_party(row[0])
        output["tvPct"][donor] = float(row[1] or 0)
        output["tvWillingness"][donor] = {}
        for dest in PIPELINE_PARTIES:
            idx = tv_col_index.get(dest)
            val = row[idx] if idx is not None and idx < len(row) else None
            output["tvWillingness"][donor][dest] = float(val) if val is not None else 0.0
    print(f"tvPct: {len(output['tvPct'])} parties")

    tvm_ws = tactical_wb["TVMatrix"]
    tvm_header = [c.value for c in next(tvm_ws.iter_rows(min_row=1, max_row=1))]
    tvm_col_index = {normalize_party(name): i for i, name in enumerate(tvm_header) if name}
    output["tvMatrix"] = {}
    for row in tvm_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        origin = normalize_party(row[0])
        output["tvMatrix"][origin] = {}
        for dest in PIPELINE_PARTIES:
            idx = tvm_col_index.get(dest)
            val = row[idx] if idx is not None and idx < len(row) else None
            output["tvMatrix"][origin][dest] = float(val) if val is not None else -1.0
    print(f"tvMatrix: {len(output['tvMatrix'])} parties")

    yg_ws = tactical_wb["Yougov"]
    yg_rows = list(yg_ws.iter_rows(min_row=2, values_only=True))
    output["yougov"] = []
    for r in yg_rows:
        if not r or r[0] is None:
            continue
        output["yougov"].append({
            "partyA": normalize_party(r[0]),
            "partyB": normalize_party(r[1]),
            "donor": normalize_party(r[2]),
            "voteFor1": float(r[3] or 0),
            "voteFor2": float(r[4] or 0),
            "voteStaying": float(r[5] or 0),
        })
    print(f"yougov: {len(output['yougov'])} rows")

    win_ws = tactical_wb["Winners"]
    output["winners"] = {}
    for row in win_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        seat, winner = row[0], row[1]
        output["winners"][seat] = normalize_party(winner) if winner else None
    print(f"winners: {len(output['winners'])} seats")

    # ---- Local flows (LocalFlows.xlsx) - per-seat overrides, applied on
    # top of the national/tactical result for the handful of seats that
    # have their own sheet. Same pattern as 05_export_svg_output.py.
    local_wb = openpyxl.load_workbook(RAW / "LocalFlows.xlsx", data_only=True)
    output["localFlows"] = {}
    for seat_name in local_wb.sheetnames:
        ws = local_wb[seat_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_index = {normalize_party(name): i for i, name in enumerate(header) if name}
        seat_matrix = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            origin = normalize_party(row[0])
            seat_matrix[origin] = {}
            for dest in PIPELINE_PARTIES:
                idx = col_index.get(dest)
                val = row[idx] if idx is not None and idx < len(row) else None
                seat_matrix[origin][dest] = float(val) if val is not None else 0.0
        output["localFlows"][seat_name] = seat_matrix
    print(f"localFlows: {len(output['localFlows'])} seats")

    WEB_DATA.mkdir(parents=True, exist_ok=True)
    out_path = WEB_DATA / "alloc.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f)

    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
